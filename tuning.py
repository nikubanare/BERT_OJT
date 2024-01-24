import random
import glob
from tqdm import tqdm

import torch
from torch.utils.data import DataLoader
from transformers import BertJapaneseTokenizer, BertForSequenceClassification
import pytorch_lightning as pl

import openpyxl

# 日本語の事前学習モデル
MODEL_NAME = 'cl-tohoku/bert-base-japanese-whole-word-masking'

# トークナイザーのロード
tokenizer = BertJapaneseTokenizer.from_pretrained(MODEL_NAME)

# 各データの形式を整える
max_length = 256
dataset_for_loader = []
bookname = 'input/merge.xlsx'
book = openpyxl.load_workbook(bookname)
sheet = book.active

max_row = sheet.max_row
for row_no in range(1,max_row+1):
    text = sheet.cell(row_no,1).value
    encoding = tokenizer(
        text,
        max_length=max_length,
        padding='max_length',
        truncation=True
    )
    encoding['labels'] = sheet.cell(row_no,2).value
    if(row_no%5==0):
        print(row_no)
    encoding = { k: torch.tensor(v) for k, v in encoding.items() }
    dataset_for_loader.append(encoding)

# データセットの分割
random.shuffle(dataset_for_loader)
n = len(dataset_for_loader)
n_train = int(0.6*n)
n_val = int(0.2*n)
dataset_train = dataset_for_loader[:n_train]
dataset_val = dataset_for_loader[n_train:n_train+n_val]
dataset_test = dataset_for_loader[n_train+n_val:]

# データセットからデータローダを作成
# 学習データはshuffle=Trueにする
dataloader_train = DataLoader(
    dataset_train, batch_size=32, shuffle=True, num_workers=8
)
dataloader_val = DataLoader(
    dataset_val, batch_size=32, num_workers=8
)
dataloader_test = DataLoader(
    dataset_test, batch_size=32,num_workers=8
)

class BertForSequenceClassification_pl(pl.LightningModule):
    
    def __init__(self, model_name, num_labels, lr):
        # model_name: Transformers のモデルの名前
        # num_labels: ラベルの数
        # lr: 学習率
        
        super().__init__()
        
        # 引数の num_labels と lr を保存
        # 例えば、self.hparams.lr で lr にアクセスできる
        # チェックポイント作成時にも自動で保存される
        self.save_hyperparameters()
        
        # BERTのロード
        self.bert_sc = BertForSequenceClassification.from_pretrained(
            model_name,
            num_labels=num_labels
        )
    
    # 学習データのミニバッチ ('batch') が与えられた時に損失を出力する関数を書く
    # batch_idx はミニバッチの番号であるが今回は使わない
    def training_step(self, batch, batch_idx):
        output = self.bert_sc(**batch)
        loss = output.loss
        self.log('train_loss', loss)
        return loss
    
    # 検証データのミニバッチが与えられたときに、
    # 検証データを評価する指標を計算する関数を書く
    def validation_step(self, batch, batch_idx):
        output = self.bert_sc(**batch)
        val_loss = output.loss
        self.log('val_loss', val_loss)
    
    # テストデータのミニバッチが与えられたときに、
    # テストデータを評価する指標を計算する関数を書く
    def test_step(self, batch, batch_idx):
        labels = batch.pop('labels')
        output = self.bert_sc(**batch)
        labels_predicted = output.logits.argmax(-1)
        num_correct = ( labels_predicted == labels ).sum().item()
        accuracy = num_correct/labels.size(0)
        self.log('accuracy', accuracy)
    
    # 学習に用いるオプティマイザを返す関数を書く
    def configure_optimizers(self):
        return torch.optim.Adam(self.parameters(), lr=self.hparams.lr)

# 学習時にモデルの重みを保存する条件を指定
checkpoint = pl.callbacks.ModelCheckpoint(
    monitor='val_loss',
    mode='min',
    save_top_k=1,
    save_weights_only=True,
    dirpath='model/',
)

# 学習の方法を指定
trainer = pl.Trainer(
    accelerator="cpu",
    max_epochs=10,
    callbacks = [checkpoint]
)

# PyTorch Lightning モデルのロード
model = BertForSequenceClassification_pl(
    MODEL_NAME, num_labels=2, lr=1e-5
)

# ファインチューニング
trainer.fit(model, dataloader_train, dataloader_val)

best_model_path = checkpoint.best_model_path
print('ベストモデルのファイル：', checkpoint.best_model_path)
print('ベストモデルの検証データに対する損失：', checkpoint.best_model_score)

test = trainer.test(dataloaders=dataloader_test)
print(f'Accuracy: {test[0]["accuracy"]:.2f}')

# PyTorch Lightning のモデルのロード
model = BertForSequenceClassification_pl.load_from_checkpoint(best_model_path)

# Transformers 対応のモデルを ./model_transformers に保存
model.bert_sc.save_pretrained('./model_transformers')
