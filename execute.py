import torch
from transformers import BertJapaneseTokenizer, AutoModelForSequenceClassification

import openpyxl

MODEL_NAME = 'cl-tohoku/bert-base-japanese-whole-word-masking'
tokenizer = BertJapaneseTokenizer.from_pretrained(MODEL_NAME)
model = AutoModelForSequenceClassification.from_pretrained('model_transformers/')

bookname = 'demo.xlsx'
book = openpyxl.load_workbook(bookname)
sheet = book.active

labels = [0,1]
model.eval()

max_row = sheet.max_row
for row_no in range(1,max_row+1):
    text = sheet.cell(row_no,1).value
    print(text)
    input = tokenizer.encode(text, return_tensors='pt')
    with torch.no_grad():
        outputs = model(input)[0]
        category = labels[torch.argmax(outputs)]
    sheet.cell(row_no,2).value = category
    print(category)
book.save('demo.xlsx')